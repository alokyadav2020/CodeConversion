import streamlit as st
import pandas as pd
import xlwings as xw
import openpyxl
from openpyxl import load_workbook
import tempfile
import os
from typing import Dict, List, Any
# import win32com.client as win32
import pythoncom

class ExcelControlsExtractor:
    def __init__(self):
        self.workbook = None
        self.xl_app = None
        
    def get_sheet_count(self, file_path: str) -> int:
        """Get the number of sheets in the Excel file using openpyxl"""
        try:
            workbook = load_workbook(file_path, read_only=True)
            sheet_count = len(workbook.sheetnames)
            workbook.close()
            return sheet_count
        except Exception as e:
            st.error(f"Error reading Excel file: {str(e)}")
            return 0
    
    def get_sheet_names(self, file_path: str) -> List[str]:
        """Get all sheet names from the Excel file"""
        try:
            workbook = load_workbook(file_path, read_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
            return sheet_names
        except Exception as e:
            st.error(f"Error getting sheet names: {str(e)}")
            return []
    
    def extract_controls_xlwings(self, file_path: str) -> Dict[str, List[Dict[str, Any]]]:
        """Extract controls using xlwings (requires Excel to be installed)"""
        controls_data = {}
        
        try:
            # Initialize xlwings app
            self.xl_app = xw.App(visible=False)
            self.workbook = xw.Book(file_path)
            
            for sheet in self.workbook.sheets:
                sheet_controls = []
                
                # Extract shapes (includes buttons, form controls)
                try:
                    for shape in sheet.shapes:
                        control_info = {
                            'type': 'Shape/Button',
                            'name': shape.name,
                            'text': getattr(shape, 'text', 'N/A'),
                            'left': getattr(shape, 'left', 'N/A'),
                            'top': getattr(shape, 'top', 'N/A'),
                            'width': getattr(shape, 'width', 'N/A'),
                            'height': getattr(shape, 'height', 'N/A')
                        }
                        sheet_controls.append(control_info)
                except Exception as e:
                    st.warning(f"Could not extract shapes from {sheet.name}: {str(e)}")
                
                # Try to extract ActiveX controls using COM
                try:
                    xl_sheet = sheet.api
                    ole_objects = xl_sheet.OLEObjects()
                    
                    for i in range(1, ole_objects.Count + 1):
                        ole_obj = ole_objects(i)
                        control_info = {
                            'type': 'ActiveX Control',
                            'name': ole_obj.Name,
                            'object_type': str(ole_obj.Object),
                            'left': ole_obj.Left,
                            'top': ole_obj.Top,
                            'width': ole_obj.Width,
                            'height': ole_obj.Height
                        }
                        
                        # Try to get specific properties based on control type
                        try:
                            if hasattr(ole_obj.Object, 'Value'):
                                control_info['value'] = ole_obj.Object.Value
                            if hasattr(ole_obj.Object, 'Caption'):
                                control_info['caption'] = ole_obj.Object.Caption
                            if hasattr(ole_obj.Object, 'Text'):
                                control_info['text'] = ole_obj.Object.Text
                        except:
                            pass
                        
                        sheet_controls.append(control_info)
                        
                except Exception as e:
                    st.warning(f"Could not extract ActiveX controls from {sheet.name}: {str(e)}")
                
                controls_data[sheet.name] = sheet_controls
                
        except Exception as e:
            st.error(f"Error extracting controls with xlwings: {str(e)}")
        finally:
            self.cleanup()
            
        return controls_data
    
    def extract_controls_openpyxl(self, file_path: str) -> Dict[str, List[Dict[str, Any]]]:
        """Extract basic information using openpyxl (limited control support)"""
        controls_data = {}
        
        try:
            workbook = load_workbook(file_path, keep_vba=True)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_controls = []
                
                # Extract data validation (dropdowns)
                if hasattr(sheet, 'data_validations') and sheet.data_validations:
                    for dv in sheet.data_validations.dataValidation:
                        for cell_range in dv.cells:
                            control_info = {
                                'type': 'Data Validation/Dropdown',
                                'cell_range': str(cell_range),
                                'validation_type': dv.type,
                                'formula1': dv.formula1,
                                'formula2': dv.formula2,
                                'allow_blank': dv.allowBlank
                            }
                            sheet_controls.append(control_info)
                
                # Extract hyperlinks
                if hasattr(sheet, '_hyperlinks'):
                    for hyperlink in sheet._hyperlinks:
                        control_info = {
                            'type': 'Hyperlink',
                            'cell': hyperlink.ref,
                            'target': hyperlink.target,
                            'display': hyperlink.display
                        }
                        sheet_controls.append(control_info)
                
                # Check for comments
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.comment:
                            control_info = {
                                'type': 'Comment',
                                'cell': cell.coordinate,
                                'text': cell.comment.text,
                                'author': cell.comment.author
                            }
                            sheet_controls.append(control_info)
                
                controls_data[sheet_name] = sheet_controls
            
            workbook.close()
            
        except Exception as e:
            st.error(f"Error extracting controls with openpyxl: {str(e)}")
            
        return controls_data
    
    def cleanup(self):
        """Clean up xlwings resources"""
        try:
            if self.workbook:
                self.workbook.close()
            if self.xl_app:
                self.xl_app.quit()
        except:
            pass

def main():
    st.set_page_config(
        page_title="Excel Controls Extractor",
        page_icon="📊",
        layout="wide"
    )
    
    st.title("📊 Excel Controls Extractor")
    st.markdown("Upload an Excel file to extract and analyze all controls (buttons, checkboxes, dropdowns, etc.)")
    
    # File upload
    uploaded_file = st.file_uploader(
        "Choose an Excel file",
        type=['xlsx', 'xlsm', 'xls'],
        help="Upload an Excel file containing form controls, ActiveX controls, or data validation"
    )
    
    if uploaded_file is not None:
        # Save uploaded file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            tmp_file.write(uploaded_file.getvalue())
            temp_file_path = tmp_file.name
        
        try:
            # Initialize extractor
            extractor = ExcelControlsExtractor()
            
            # Display file info
            st.success(f"✅ File uploaded: {uploaded_file.name}")
            
            # Get basic file information
            sheet_count = extractor.get_sheet_count(temp_file_path)
            sheet_names = extractor.get_sheet_names(temp_file_path)
            
            col1, col2 = st.columns(2)
            with col1:
                st.metric("Number of Sheets", sheet_count)
            with col2:
                st.metric("File Size", f"{uploaded_file.size / 1024:.1f} KB")
            
            if sheet_names:
                st.subheader("📋 Sheet Names")
                for i, name in enumerate(sheet_names, 1):
                    st.write(f"{i}. {name}")
            
            # Extraction method selection
            st.subheader("🔧 Control Extraction Method")
            extraction_method = st.radio(
                "Select extraction method:",
                ["Basic (openpyxl)", "Advanced (xlwings - requires Excel)"],
                help="Basic method works without Excel installed but has limited control detection. Advanced method requires Excel but can detect more control types."
            )
            
            if st.button("🔍 Extract Controls", type="primary"):
                with st.spinner("Extracting controls..."):
                    if extraction_method == "Basic (openpyxl)":
                        controls_data = extractor.extract_controls_openpyxl(temp_file_path)
                    else:
                        controls_data = extractor.extract_controls_xlwings(temp_file_path)
                
                # Display results
                st.subheader("📊 Extracted Controls")
                
                if not any(controls_data.values()):
                    st.info("No controls found in the Excel file.")
                else:
                    # Summary
                    total_controls = sum(len(controls) for controls in controls_data.values())
                    st.metric("Total Controls Found", total_controls)
                    
                    # Display controls by sheet
                    for sheet_name, controls in controls_data.items():
                        if controls:
                            st.subheader(f"Sheet: {sheet_name}")
                            
                            # Create DataFrame for better display
                            df = pd.DataFrame(controls)
                            
                            # Display as expandable table
                            with st.expander(f"View {len(controls)} controls in '{sheet_name}'", expanded=True):
                                st.dataframe(df, use_container_width=True)
                            
                            # Control type summary
                            if 'type' in df.columns:
                                control_types = df['type'].value_counts()
                                st.write("**Control Types:**")
                                for control_type, count in control_types.items():
                                    st.write(f"- {control_type}: {count}")
                        else:
                            st.write(f"**Sheet: {sheet_name}** - No controls found")
                
                # Export option
                if any(controls_data.values()):
                    st.subheader("💾 Export Results")
                    
                    # Prepare data for export
                    export_data = []
                    for sheet_name, controls in controls_data.items():
                        for control in controls:
                            control['sheet_name'] = sheet_name
                            export_data.append(control)
                    
                    if export_data:
                        export_df = pd.DataFrame(export_data)
                        csv = export_df.to_csv(index=False)
                        
                        st.download_button(
                            label="📄 Download Controls as CSV",
                            data=csv,
                            file_name=f"{uploaded_file.name}_controls.csv",
                            mime="text/csv"
                        )
        
        except Exception as e:
            st.error(f"Error processing file: {str(e)}")
        
        finally:
            # Clean up temporary file
            try:
                os.unlink(temp_file_path)
            except:
                pass
    
    # Instructions
    with st.sidebar:
        st.header("📖 Instructions")
        st.markdown("""
        **How to use:**
        1. Upload an Excel file (.xlsx, .xlsm, .xls)
        2. Choose extraction method
        3. Click "Extract Controls"
        4. View and export results
        
        **Supported Controls:**
        - **Basic Method:**
          - Data validation (dropdowns)
          - Hyperlinks
          - Comments
        
        - **Advanced Method:**
          - Form controls (buttons, checkboxes)
          - ActiveX controls
          - Shapes
          - All basic method controls
        
        **Requirements:**
        - Basic method: Only Python packages
        - Advanced method: Microsoft Excel installed
        """)
        
        # st.header("🔧 Dependencies")
#         st.code("""
# pip install streamlit
# pip install pandas
# pip install openpyxl
# pip install xlwings
# pip install pywin32
#         """)

if __name__ == "__main__":
    main()