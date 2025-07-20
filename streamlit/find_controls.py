import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import io

def get_cell_address(anchor):
    """
    Converts an openpyxl anchor object to a human-readable cell address like 'A1'.
    """
    # The anchor gives row and column numbers (0-indexed)
    col_letter = get_column_letter(anchor.from_col + 1)
    row_num = anchor.from_row + 1
    return f"{col_letter}{row_num}"

def extract_all_controls(file_bytes):
    """
    Identifies all sheets in an Excel file and extracts all controls from each sheet.

    Args:
        file_bytes (bytes): The content of the uploaded Excel file.

    Returns:
        tuple: A tuple containing:
               - list: A list of dictionaries, where each dict contains control details.
               - list: A list of all sheet names found in the workbook.
    """
    all_controls = []
    
    try:
        # Load the workbook from the in-memory file bytes
        workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        sheet_names = workbook.sheetnames

        # A mapping for common Form Control tags to human-readable names
        form_control_map = {
            'sp': 'Spinner or Scrollbar',
            'btn': 'Button',
            'chx': 'Checkbox',
            'drop': 'Dropdown (Combo Box)',
            'opt': 'Option Button (Radio Button)',
            'gbox': 'Group Box',
            'lbl': 'Label',
            'edit': 'Text Box' # Less common, usually an ActiveX control
        }
        
        # A mapping for common ActiveX control ProgIDs
        activex_control_map = {
            'Forms.CommandButton.1': 'ActiveX Command Button',
            'Forms.CheckBox.1': 'ActiveX CheckBox',
            'Forms.ComboBox.1': 'ActiveX ComboBox (Dropdown)',
            'Forms.ListBox.1': 'ActiveX ListBox',
            'Forms.TextBox.1': 'ActiveX TextBox',
            'Forms.OptionButton.1': 'ActiveX Option Button',
            'Forms.ToggleButton.1': 'ActiveX Toggle Button',
            'Forms.Frame.1': 'ActiveX Frame',
            'Forms.Label.1': 'ActiveX Label',
            'Forms.ScrollBar.1': 'ActiveX ScrollBar',
            'Forms.SpinButton.1': 'ActiveX SpinButton',
        }

        # --- Loop through each sheet to find controls ---
        for sheet_name in sheet_names:
            ws = workbook[sheet_name]

            # 1. Extract FORM CONTROLS
            # These are stored in a "private" attribute `_controls`
            if hasattr(ws, '_controls') and ws._controls:
                for control in ws._controls:
                    control_type_tag = control.tagname
                    control_details = {
                        'Sheet Name': sheet_name,
                        'Control Name': control.name,
                        'Control Type': f"Form Control - {form_control_map.get(control_type_tag, 'Unknown')}",
                        'Location (Top-Left Cell)': get_cell_address(control.anchor)
                    }
                    all_controls.append(control_details)

            # 2. Extract ACTIVEX CONTROLS
            # These are stored in the `ole_objects` attribute
            if hasattr(ws, 'ole_objects') and ws.ole_objects:
                for ole_obj in ws.ole_objects:
                    # Check if it's a known control type via its ProgID
                    if ole_obj.progId and ole_obj.progId in activex_control_map:
                        control_details = {
                            'Sheet Name': sheet_name,
                            'Control Name': ole_obj.name,
                            'Control Type': activex_control_map.get(ole_obj.progId, 'Unknown ActiveX'),
                            'Location (Top-Left Cell)': get_cell_address(ole_obj.anchor)
                        }
                        all_controls.append(control_details)
                        
        return all_controls, sheet_names

    except Exception as e:
        st.error(f"An error occurred while processing the Excel file: {e}")
        st.error("Please ensure you have uploaded a valid .xlsx file.")
        return [], []

# --- Main Streamlit App Logic ---
def main():
    st.set_page_config(layout="wide")
    st.title("📊 Excel Control Extractor")
    st.markdown("""
    Upload an Excel file (`.xlsx`) to identify and list all Form Controls and ActiveX Controls 
    (like buttons, checkboxes, dropdowns, etc.) from every sheet.
    """)

    # File uploader widget
    uploaded_file = st.file_uploader("Choose an Excel file", type=['xlsx', 'xlsm', 'xls'])

    if uploaded_file is not None:
        # To read file as bytes:
        file_bytes = uploaded_file.getvalue()
        
        with st.spinner('Analyzing your Excel file... This may take a moment.'):
            controls_found, sheet_names = extract_all_controls(file_bytes)

        st.header("Extraction Results")

        if not sheet_names:
            # This handles the case where the file was invalid and the function returned empty lists
            st.warning("Could not process the file.")
        else:
            # Show summary information
            col1, col2 = st.columns(2)
            col1.metric("Total Sheets Found", len(sheet_names))
            col2.metric("Total Controls Found", len(controls_found))
            
            st.info(f"**Sheets Scanned:** `{'`, `'.join(sheet_names)}`")

            if controls_found:
                # Display the results in a clean table (DataFrame)
                st.subheader("List of All Controls")
                df = pd.DataFrame(controls_found)
                st.dataframe(df, use_container_width=True)
            else:
                st.success("✅ The file was processed successfully, but no controls were found.")

    else:
        st.info("Awaiting for an Excel file to be uploaded.")
if __name__ == "__main__":
    main()